from __future__ import annotations

import logging
import sys
import traceback
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
LOG_DIR = BASE_DIR / "log"

LOG_FILE = LOG_DIR / "data_cleaner.log"

REQUIRED_COLUMNS = [
    "customer_id",
    "customer_name",
    "email",
    "phone",
    "country",
    "signup_date",
    "product",
    "quantity",
    "unit_price",
    "total_amount",
    "status",
]

TEXT_COLUMNS = [
    "customer_id",
    "customer_name",
    "email",
    "phone",
    "country",
    "product",
    "status",
]

NUMERIC_COLUMNS = [
    "quantity",
    "unit_price",
    "total_amount",
]

KEY_FIELDS = [
    "customer_id",
    "customer_name",
    "product",
]

VALID_STATUSES = {
    "active",
    "inactive",
    "pending",
    "cancelled",
    "completed",
    "returned",
    "processing",
}

STATUS_NORMALIZATION_MAP = {
    "active": "Active",
    "inactive": "Inactive",
    "pending": "Pending",
    "cancelled": "Cancelled",
    "canceled": "Cancelled",
    "completed": "Completed",
    "returned": "Returned",
    "processing": "Processing",
    "process": "Processing",
}


@dataclass
class QualityMetrics:
    source_file: str
    raw_row_count: int = 0
    row_count_after_empty_row_removal: int = 0
    duplicate_rows_removed: int = 0
    missing_key_field_rows_flagged: int = 0
    invalid_date_count: int = 0
    invalid_numeric_count: int = 0
    negative_numeric_count: int = 0
    total_recalculated_count: int = 0
    invalid_status_count: int = 0
    final_clean_row_count: int = 0
    final_issue_row_count: int = 0

    def to_summary_rows(self) -> list[list[Any]]:
        return [
            ["Source File", self.source_file],
            ["Raw Row Count", self.raw_row_count],
            ["After Empty Row Removal", self.row_count_after_empty_row_removal],
            ["Duplicate Rows Removed", self.duplicate_rows_removed],
            ["Rows Flagged - Missing Key Fields", self.missing_key_field_rows_flagged],
            ["Invalid Date Count", self.invalid_date_count],
            ["Invalid Numeric Count", self.invalid_numeric_count],
            ["Negative Numeric Count", self.negative_numeric_count],
            ["Totals Recalculated", self.total_recalculated_count],
            ["Invalid Status Count", self.invalid_status_count],
            ["Final Clean Row Count", self.final_clean_row_count],
            ["Final Issue Row Count", self.final_issue_row_count],
            ["Run Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]


def setup_directories() -> None:
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)


def setup_logging() -> None:
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    if logger.handlers:
        logger.handlers.clear()

    formatter = logging.Formatter(
        fmt="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    file_handler = logging.FileHandler(LOG_FILE, encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)

    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setLevel(logging.INFO)
    stream_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)


def standardize_column_name(column_name: Any) -> str:
    column = str(column_name).strip().lower()
    column = column.replace(" ", "_")
    column = column.replace("-", "_")
    column = column.replace("/", "_")
    while "__" in column:
        column = column.replace("__", "_")
    return column


def find_input_file() -> Path:
    supported_patterns = ["*.csv", "*.xlsx", "*.xls"]
    files: list[Path] = []

    for pattern in supported_patterns:
        files.extend(INPUT_DIR.glob(pattern))

    files = sorted(files)

    if not files:
        raise FileNotFoundError(
            f"No input file found in {INPUT_DIR}. "
            "Add a CSV or Excel file to the input folder."
        )

    selected_file = files[0]
    logging.info("Selected input file: %s", selected_file.name)
    return selected_file


def load_input_data(file_path: Path) -> pd.DataFrame:
    logging.info("Loading input data from %s", file_path.name)

    if file_path.suffix.lower() == ".csv":
        df = pd.read_csv(file_path, dtype=str)
    elif file_path.suffix.lower() in {".xlsx", ".xls"}:
        df = pd.read_excel(file_path, dtype=str)
    else:
        raise ValueError(f"Unsupported file type: {file_path.suffix}")

    logging.info("Loaded %s raw rows and %s columns", len(df), len(df.columns))
    return df


def validate_required_columns(df: pd.DataFrame) -> None:
    standardized_columns = [standardize_column_name(col) for col in df.columns]
    missing_columns = [col for col in REQUIRED_COLUMNS if col not in standardized_columns]

    if missing_columns:
        raise ValueError(
            "Missing required columns: " + ", ".join(missing_columns)
        )


def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [standardize_column_name(col) for col in df.columns]
    return df


def replace_placeholder_missing_values(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    missing_markers = {
        "": pd.NA,
        " ": pd.NA,
        "  ": pd.NA,
        "   ": pd.NA,
        "nan": pd.NA,
        "none": pd.NA,
        "null": pd.NA,
        "n/a": pd.NA,
        "na": pd.NA,
        "-": pd.NA,
        "--": pd.NA,
    }

    for column in df.columns:
        df[column] = df[column].apply(
            lambda value: str(value).strip() if pd.notna(value) else value
        )
        df[column] = df[column].replace(missing_markers)

    return df


def remove_fully_empty_rows(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    before_count = len(df)
    cleaned_df = df.dropna(how="all").copy()
    removed = before_count - len(cleaned_df)
    logging.info("Removed %s fully empty rows", removed)
    return cleaned_df, removed


def clean_text_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    for column in TEXT_COLUMNS:
        if column not in df.columns:
            continue

        df[column] = df[column].apply(
            lambda value: str(value).strip() if pd.notna(value) else value
        )

    if "customer_name" in df.columns:
        df["customer_name"] = df["customer_name"].apply(
            lambda value: value.title() if pd.notna(value) else value
        )

    if "country" in df.columns:
        df["country"] = df["country"].apply(
            lambda value: value.title() if pd.notna(value) else value
        )

    if "product" in df.columns:
        df["product"] = df["product"].apply(
            lambda value: value.title() if pd.notna(value) else value
        )

    if "email" in df.columns:
        df["email"] = df["email"].apply(
            lambda value: value.lower() if pd.notna(value) else value
        )

    if "status" in df.columns:
        df["status"] = df["status"].apply(
            lambda value: STATUS_NORMALIZATION_MAP.get(value.lower(), value.title())
            if pd.notna(value)
            else value
        )

    if "phone" in df.columns:
        df["phone"] = df["phone"].apply(normalize_phone)

    return df


def normalize_phone(value: Any) -> Any:
    if pd.isna(value):
        return value

    raw = str(value).strip()
    raw = raw.replace("(", "").replace(")", "")
    raw = raw.replace("-", "").replace(".", "").replace(" ", "")

    return raw if raw else pd.NA


def clean_customer_id(value: Any) -> Any:
    if pd.isna(value):
        return value

    customer_id = str(value).strip().upper().replace(" ", "")
    return customer_id if customer_id else pd.NA


def clean_status_column(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    df = df.copy()
    invalid_status_count = 0

    if "status" not in df.columns:
        return df, invalid_status_count

    normalized_status_series = df["status"].apply(
        lambda value: value.strip().lower() if pd.notna(value) else value
    )

    invalid_status_mask = normalized_status_series.notna() & ~normalized_status_series.isin(
        VALID_STATUSES
    )
    invalid_status_count = int(invalid_status_mask.sum())

    if invalid_status_count:
        logging.info("Found %s invalid status values", invalid_status_count)

    df.loc[invalid_status_mask, "status"] = "Review Required"

    return df, invalid_status_count


def parse_date_series(series: pd.Series) -> tuple[pd.Series, int]:
    parsed_dates = pd.to_datetime(series, errors="coerce", dayfirst=False)
    invalid_count = int(series.notna().sum() - parsed_dates.notna().sum())

    formatted_dates = parsed_dates.dt.strftime("%Y-%m-%d")
    formatted_dates = formatted_dates.where(parsed_dates.notna(), pd.NA)

    return formatted_dates, invalid_count


def clean_date_columns(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    df = df.copy()
    invalid_date_count = 0

    if "signup_date" in df.columns:
        df["signup_date"], invalid_date_count = parse_date_series(df["signup_date"])
        logging.info("Invalid dates identified: %s", invalid_date_count)

    return df, invalid_date_count


def clean_numeric_value(value: Any) -> Any:
    if pd.isna(value):
        return pd.NA

    text = str(value).strip()

    if not text:
        return pd.NA

    text = text.replace("$", "")
    text = text.replace("€", "")
    text = text.replace("£", "")
    text = text.replace(",", "")
    text = text.replace(" ", "")

    if text.endswith(".0.0"):
        text = text[:-2]

    try:
        return float(text)
    except ValueError:
        return pd.NA


def clean_numeric_columns(df: pd.DataFrame) -> tuple[pd.DataFrame, int, int]:
    df = df.copy()
    invalid_numeric_count = 0
    negative_numeric_count = 0

    for column in NUMERIC_COLUMNS:
        if column not in df.columns:
            continue

        original_non_null = int(df[column].notna().sum())

        df[column] = df[column].apply(clean_numeric_value)

        cleaned_non_null = int(df[column].notna().sum())
        column_invalid = original_non_null - cleaned_non_null
        invalid_numeric_count += max(column_invalid, 0)

        negative_count = int((df[column] < 0).fillna(False).sum())
        negative_numeric_count += negative_count

        logging.info(
            "Numeric cleanup for %s | invalid: %s | negative: %s",
            column,
            column_invalid,
            negative_count,
        )

    return df, invalid_numeric_count, negative_numeric_count


def normalize_identifiers(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    if "customer_id" in df.columns:
        df["customer_id"] = df["customer_id"].apply(clean_customer_id)

    return df


def remove_duplicates(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    before_count = len(df)
    deduplicated_df = df.drop_duplicates().copy()
    removed = before_count - len(deduplicated_df)
    logging.info("Removed %s duplicate rows", removed)
    return deduplicated_df, removed


def recalculate_total_amount(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    df = df.copy()
    recalculated_count = 0

    if not {"quantity", "unit_price", "total_amount"}.issubset(df.columns):
        return df, recalculated_count

    quantity_numeric = pd.to_numeric(df["quantity"], errors="coerce")
    unit_price_numeric = pd.to_numeric(df["unit_price"], errors="coerce")
    total_numeric = pd.to_numeric(df["total_amount"], errors="coerce")

    calculated_total = quantity_numeric * unit_price_numeric

    mismatch_mask = (
        quantity_numeric.notna()
        & unit_price_numeric.notna()
        & (
            total_numeric.isna()
            | (total_numeric.round(2) != calculated_total.round(2))
        )
    )

    recalculated_count = int(mismatch_mask.sum())

    df.loc[mismatch_mask, "total_amount"] = calculated_total[mismatch_mask].round(2)

    logging.info("Recalculated total_amount for %s rows", recalculated_count)
    return df, recalculated_count


def build_issue_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    issue_columns = [
        "issue_missing_key_field",
        "issue_invalid_signup_date",
        "issue_invalid_numeric_data",
        "issue_negative_values",
        "issue_status_review_required",
        "issue_notes",
        "record_status",
    ]

    for column in issue_columns:
        if column not in df.columns:
            df[column] = pd.NA

    missing_key_mask = pd.Series(False, index=df.index)
    for field in KEY_FIELDS:
        if field in df.columns:
            missing_key_mask = missing_key_mask | df[field].isna()

    invalid_date_mask = df["signup_date"].isna() if "signup_date" in df.columns else pd.Series(False, index=df.index)

    invalid_numeric_mask = pd.Series(False, index=df.index)
    for field in ["quantity", "unit_price", "total_amount"]:
        if field in df.columns:
            invalid_numeric_mask = invalid_numeric_mask | df[field].isna()

    negative_numeric_mask = pd.Series(False, index=df.index)
    for field in ["quantity", "unit_price", "total_amount"]:
        if field in df.columns:
            negative_numeric_mask = negative_numeric_mask | (pd.to_numeric(df[field], errors="coerce") < 0).fillna(False)

    status_review_mask = (
        df["status"].eq("Review Required")
        if "status" in df.columns
        else pd.Series(False, index=df.index)
    )

    df["issue_missing_key_field"] = missing_key_mask.map({True: "Yes", False: "No"})
    df["issue_invalid_signup_date"] = invalid_date_mask.map({True: "Yes", False: "No"})
    df["issue_invalid_numeric_data"] = invalid_numeric_mask.map({True: "Yes", False: "No"})
    df["issue_negative_values"] = negative_numeric_mask.map({True: "Yes", False: "No"})
    df["issue_status_review_required"] = status_review_mask.map({True: "Yes", False: "No"})

    issue_notes = []

    for _, row in df.iterrows():
        notes: list[str] = []

        if row["issue_missing_key_field"] == "Yes":
            notes.append("Missing required business field")
        if row["issue_invalid_signup_date"] == "Yes":
            notes.append("Invalid or missing signup date")
        if row["issue_invalid_numeric_data"] == "Yes":
            notes.append("Invalid or missing numeric value")
        if row["issue_negative_values"] == "Yes":
            notes.append("Negative numeric value detected")
        if row["issue_status_review_required"] == "Yes":
            notes.append("Unexpected status value")

        issue_notes.append("; ".join(notes) if notes else "No issues detected")

    df["issue_notes"] = issue_notes

    issue_flag_mask = (
        missing_key_mask
        | invalid_date_mask
        | invalid_numeric_mask
        | negative_numeric_mask
        | status_review_mask
    )

    df["record_status"] = issue_flag_mask.map({True: "Flagged", False: "Clean"})

    return df


def split_clean_and_issue_data(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    clean_df = df[df["record_status"] == "Clean"].copy()
    issue_df = df[df["record_status"] == "Flagged"].copy()
    return clean_df, issue_df


def build_issue_summary_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    summary_rows = [
        ["Missing Key Fields", int((df["issue_missing_key_field"] == "Yes").sum())],
        ["Invalid Signup Dates", int((df["issue_invalid_signup_date"] == "Yes").sum())],
        ["Invalid Numeric Data", int((df["issue_invalid_numeric_data"] == "Yes").sum())],
        ["Negative Numeric Values", int((df["issue_negative_values"] == "Yes").sum())],
        ["Unexpected Status Values", int((df["issue_status_review_required"] == "Yes").sum())],
        ["Flagged Records", int((df["record_status"] == "Flagged").sum())],
        ["Clean Records", int((df["record_status"] == "Clean").sum())],
    ]

    return pd.DataFrame(summary_rows, columns=["Metric", "Value"])


def sort_final_output(df: pd.DataFrame) -> pd.DataFrame:
    preferred_order = [
        "customer_id",
        "customer_name",
        "email",
        "phone",
        "country",
        "signup_date",
        "product",
        "quantity",
        "unit_price",
        "total_amount",
        "status",
        "record_status",
        "issue_missing_key_field",
        "issue_invalid_signup_date",
        "issue_invalid_numeric_data",
        "issue_negative_values",
        "issue_status_review_required",
        "issue_notes",
    ]

    available_columns = [col for col in preferred_order if col in df.columns]
    remaining_columns = [col for col in df.columns if col not in available_columns]

    return df[available_columns + remaining_columns].copy()


def write_summary_text_file(
    metrics: QualityMetrics,
    summary_df: pd.DataFrame,
    issue_summary_df: pd.DataFrame,
    output_path: Path,
) -> None:
    lines: list[str] = []
    lines.append("CSV and Excel Data Cleaner Automation - Summary Report")
    lines.append("=" * 60)
    lines.append("")

    for label, value in metrics.to_summary_rows():
        lines.append(f"{label}: {value}")

    lines.append("")
    lines.append("Business Data Quality Summary")
    lines.append("-" * 60)

    for _, row in summary_df.iterrows():
        lines.append(f"{row['Metric']}: {row['Value']}")

    lines.append("")
    lines.append("Issue Summary")
    lines.append("-" * 60)

    for _, row in issue_summary_df.iterrows():
        lines.append(f"{row['Metric']}: {row['Value']}")

    output_path.write_text("\n".join(lines), encoding="utf-8")
    logging.info("Created summary text report: %s", output_path.name)


def apply_excel_styling(workbook: Workbook) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=12)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top")

                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value)) if cell.value is not None else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    continue

            adjusted_width = min(max(max_length + 2, 12), 40)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        worksheet.sheet_view.showGridLines = True

        if worksheet.title in {"Data Quality Summary", "Issue Summary"}:
            for cell in worksheet["A"]:
                cell.font = title_font if cell.row > 1 else header_font


def write_excel_report(
    clean_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    issue_summary_df: pd.DataFrame,
    output_path: Path,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        clean_df.to_excel(writer, index=False, sheet_name="Cleaned Data")
        summary_df.to_excel(writer, index=False, sheet_name="Data Quality Summary")
        issue_summary_df.to_excel(writer, index=False, sheet_name="Issue Summary")

        workbook = writer.book
        apply_excel_styling(workbook)

    logging.info("Created formatted Excel report: %s", output_path.name)


def prepare_quality_summary_dataframe(metrics: QualityMetrics) -> pd.DataFrame:
    return pd.DataFrame(metrics.to_summary_rows(), columns=["Metric", "Value"])


def process_data(file_path: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, QualityMetrics]:
    metrics = QualityMetrics(source_file=file_path.name)

    df = load_input_data(file_path)
    metrics.raw_row_count = len(df)

    validate_required_columns(df)
    df = standardize_columns(df)
    df = replace_placeholder_missing_values(df)

    df, empty_rows_removed = remove_fully_empty_rows(df)
    metrics.row_count_after_empty_row_removal = len(df)

    df = clean_text_columns(df)
    df = normalize_identifiers(df)

    df, invalid_status_count = clean_status_column(df)
    metrics.invalid_status_count = invalid_status_count

    df, invalid_date_count = clean_date_columns(df)
    metrics.invalid_date_count = invalid_date_count

    df, invalid_numeric_count, negative_numeric_count = clean_numeric_columns(df)
    metrics.invalid_numeric_count = invalid_numeric_count
    metrics.negative_numeric_count = negative_numeric_count

    df, duplicate_removed = remove_duplicates(df)
    metrics.duplicate_rows_removed = duplicate_removed

    df, recalculated_count = recalculate_total_amount(df)
    metrics.total_recalculated_count = recalculated_count

    df = build_issue_columns(df)

    metrics.missing_key_field_rows_flagged = int(
        (df["issue_missing_key_field"] == "Yes").sum()
    )

    clean_df, issue_df = split_clean_and_issue_data(df)

    metrics.final_clean_row_count = len(clean_df)
    metrics.final_issue_row_count = len(issue_df)

    final_df = sort_final_output(df)
    issue_summary_df = build_issue_summary_dataframe(final_df)
    quality_summary_df = prepare_quality_summary_dataframe(metrics)

    return final_df, quality_summary_df, issue_summary_df, metrics

def fill_unknown_values(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    return df.fillna("[Unknown]")

def export_outputs(
    final_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    issue_summary_df: pd.DataFrame,
) -> None:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    excel_output_path = OUTPUT_DIR / f"cleaned_business_data_{timestamp}.xlsx"
    csv_output_path = OUTPUT_DIR / f"cleaned_business_data_{timestamp}.csv"
    summary_output_path = OUTPUT_DIR / f"data_quality_summary_{timestamp}.txt"

    clean_export_df = final_df.copy()

    clean_export_df = final_df.copy()
    clean_export_df = fill_unknown_values(clean_export_df)
    write_excel_report(clean_export_df, summary_df, issue_summary_df, excel_output_path)

    clean_export_df.to_csv(csv_output_path, index=False, encoding="utf-8")
    logging.info("Created cleaned CSV file: %s", csv_output_path.name)

    metrics_proxy = QualityMetrics(source_file="Included in workbook")
    for _, row in summary_df.iterrows():
        metric = row["Metric"]
        value = row["Value"]

        mapping = {
            "Source File": "source_file",
            "Raw Row Count": "raw_row_count",
            "After Empty Row Removal": "row_count_after_empty_row_removal",
            "Duplicate Rows Removed": "duplicate_rows_removed",
            "Rows Flagged - Missing Key Fields": "missing_key_field_rows_flagged",
            "Invalid Date Count": "invalid_date_count",
            "Invalid Numeric Count": "invalid_numeric_count",
            "Negative Numeric Count": "negative_numeric_count",
            "Totals Recalculated": "total_recalculated_count",
            "Invalid Status Count": "invalid_status_count",
            "Final Clean Row Count": "final_clean_row_count",
            "Final Issue Row Count": "final_issue_row_count",
        }

        if metric in mapping:
            setattr(metrics_proxy, mapping[metric], value)

    write_summary_text_file(
        metrics=metrics_proxy,
        summary_df=summary_df,
        issue_summary_df=issue_summary_df,
        output_path=summary_output_path,
    )


def log_metrics(metrics: QualityMetrics) -> None:
    logging.info("Processing completed successfully")
    logging.info("Raw row count: %s", metrics.raw_row_count)
    logging.info("Rows after empty-row removal: %s", metrics.row_count_after_empty_row_removal)
    logging.info("Duplicate rows removed: %s", metrics.duplicate_rows_removed)
    logging.info("Missing key field rows flagged: %s", metrics.missing_key_field_rows_flagged)
    logging.info("Invalid date count: %s", metrics.invalid_date_count)
    logging.info("Invalid numeric count: %s", metrics.invalid_numeric_count)
    logging.info("Negative numeric count: %s", metrics.negative_numeric_count)
    logging.info("Totals recalculated: %s", metrics.total_recalculated_count)
    logging.info("Invalid status count: %s", metrics.invalid_status_count)
    logging.info("Final clean row count: %s", metrics.final_clean_row_count)
    logging.info("Final issue row count: %s", metrics.final_issue_row_count)


def main() -> None:
    setup_directories()
    setup_logging()

    logging.info("Starting CSV and Excel data cleaner automation")

    try:
        input_file = find_input_file()
        final_df, summary_df, issue_summary_df, metrics = process_data(input_file)
        export_outputs(final_df, summary_df, issue_summary_df)
        log_metrics(metrics)

        logging.info("All output files created in: %s", OUTPUT_DIR)

    except Exception as exc:
        logging.error("Automation failed: %s", exc)
        logging.error("Traceback:\n%s", traceback.format_exc())
        raise


if __name__ == "__main__":
    main()